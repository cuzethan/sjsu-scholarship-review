"""
Lambda: parse-applications  (Phase 1 — SJSU General only)

Triggered by S3 event when a new .xlsx lands in the data/ prefix. Parses SJSU
General Scholarship workbooks (25-26 and 26-27) into normalized application
records and writes them to DynamoDB keyed by a deterministic `application_key`.

Uses openpyxl directly (no pandas/numpy) so the Lambda package stays tiny.

Phase-1 scope:
- Only SJSU General workbooks are parsed; specialized workbooks are skipped and
  logged as unsupported.
- No rubric_id in the schema (single shared rubric handled downstream).
- Records written with status="parsed"; a DynamoDB Stream triggers scoring.
"""

import io
import json
import logging
import os
from datetime import datetime, timezone
from urllib.parse import unquote_plus

import boto3
import openpyxl

from scholarship_config import (
    NUMERIC_FIELDS, SCHOLARSHIP_SCOPE, build_application_key,
    extract_year, identify_scholarship,
)

logger = logging.getLogger()
logger.setLevel(logging.INFO)

APPLICATIONS_TABLE = os.environ.get("APPLICATIONS_TABLE", "sjsu-applications")
AWS_REGION = os.environ.get("AWS_REGION", "us-west-2")


def get_dynamo_table():
    dynamodb = boto3.resource("dynamodb", region_name=AWS_REGION)
    return dynamodb.Table(APPLICATIONS_TABLE)


def read_xlsx_from_s3(bucket: str, key: str):
    """Download an xlsx from S3 and yield (headers, rows, sheet_name).

    rows is a generator of dicts {column_header: value} using the first row as
    headers. Uses openpyxl read-only for low memory.
    """
    s3 = boto3.client("s3", region_name=AWS_REGION)
    body = s3.get_object(Bucket=bucket, Key=key)["Body"].read()
    wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], iter([]), ws.title
    headers = [str(h).strip() if h is not None else "" for h in header]

    def row_dicts():
        for raw in rows_iter:
            yield {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}

    return headers, row_dicts(), ws.title


def clean_value(value, is_numeric: bool = False):
    """Clean a cell value: handle None/blank, strip strings, avoid float artifacts."""
    if value is None:
        return None
    if is_numeric:
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    if isinstance(value, float):
        return str(value) if value != int(value) else str(int(value))
    if isinstance(value, str):
        return value.strip() or None
    return str(value)


def normalize_row(row: dict, config: dict, year: str,
                  file_name: str, sheet_name: str, row_number: int) -> dict | None:
    """Convert one row-dict into a phase-1 SJSU General application dict."""
    column_map = config["column_map"]
    essay_fields = config["essay_fields"]

    record = {
        "scholarship_scope": SCHOLARSHIP_SCOPE,
        "year": year,
        "student_uuid": None,
        "availability_id": None,
        "student_name": None,     # not present in this anonymized data
        "gpa": None,
        "academic_program": None,
        "academic_level": None,
        "major": None,
        "qa_pairs": [],
        "source": {"file_name": file_name, "sheet_name": sheet_name, "row_number": row_number},
        "status": "parsed",
    }

    for raw_col, field_name in column_map.items():
        if raw_col not in row:
            continue
        if field_name == "availability_id":
            v = row[raw_col]
            record["availability_id"] = str(v).strip() if v is not None else None
        else:
            record[field_name] = clean_value(row[raw_col], is_numeric=field_name in NUMERIC_FIELDS)

    if not record["student_uuid"]:
        return None

    record["application_key"] = build_application_key(record["student_uuid"])

    for essay_def in essay_fields:
        answer = None
        for col in [essay_def["raw_column"]] + essay_def.get("alt_columns", []):
            if col in row:
                answer = clean_value(row[col])
                if answer is not None:
                    break
        if answer is None:
            continue
        qa = {"question_id": essay_def["question_id"],
              "question": essay_def["question"], "answer": answer}
        if "topic" in essay_def:
            qa["topic"] = essay_def["topic"]
        record["qa_pairs"].append(qa)

    return record


def parse_file(bucket: str, key: str) -> list[dict]:
    """Parse a single xlsx file. Only SJSU General workbooks are processed."""
    filename = key.split("/")[-1]

    config = identify_scholarship(filename)
    if config is None:
        logger.info(f"UNSUPPORTED in phase 1 (SJSU General only), skipping: '{filename}'")
        return []

    year = extract_year(filename) or "unknown"
    if year != "26-27":
        logger.info(f"Phase 1 uses 26-27 data only; skipping year '{year}': '{filename}'")
        return []
    _, rows, sheet_name = read_xlsx_from_s3(bucket, key)
    logger.info(f"Parsing SJSU General | {filename} | year {year}")

    records = []
    for idx, row in enumerate(rows):
        rec = normalize_row(row, config, year, file_name=filename,
                            sheet_name=sheet_name, row_number=idx + 2)
        if rec is not None:
            records.append(rec)
    logger.info(f"Parsed {len(records)} SJSU General applications from {filename}")
    return records


def write_to_dynamodb(records: list[dict], source_file: str) -> int:
    """Batch write records keyed by application_key (deterministic, idempotent)."""
    table = get_dynamo_table()
    parsed_at = datetime.now(timezone.utc).isoformat()
    written = 0

    with table.batch_writer(overwrite_by_pkeys=["application_key"]) as batch:
        for record in records:
            if not record.get("application_key"):
                continue
            item = {
                "application_key": record["application_key"],
                "scholarship_scope": record["scholarship_scope"],
                "year": record["year"],
                "status": record.get("status", "parsed"),
                "source_file": source_file,
                "parsed_at": parsed_at,
            }
            for field in ("availability_id", "student_name",
                          "academic_program", "academic_level", "major"):
                if record.get(field) is not None:
                    item[field] = record[field]
            if record.get("gpa") is not None:
                item["gpa"] = str(record["gpa"])
            if record.get("qa_pairs"):
                item["qa_pairs"] = record["qa_pairs"]
            if record.get("source"):
                item["source"] = record["source"]
            batch.put_item(Item=item)
            written += 1

    logger.info(f"Wrote {written} records to {APPLICATIONS_TABLE}")
    return written


def handler(event, context):
    """S3-event entry point."""
    logger.info(f"Event received: {json.dumps(event)}")
    records_written = 0

    for record in event.get("Records", []):
        bucket = record["s3"]["bucket"]["name"]
        key = unquote_plus(record["s3"]["object"]["key"])

        if not key.startswith("data/") or not key.endswith(".xlsx"):
            logger.info(f"Skipping non-xlsx or non-data/ file: {key}")
            continue
        if key.split("/")[-1].startswith("~$"):
            logger.info(f"Skipping temp file: {key}")
            continue

        logger.info(f"Processing: s3://{bucket}/{key}")
        try:
            applications = parse_file(bucket, key)
            if applications:
                records_written += write_to_dynamodb(applications, source_file=key)
            else:
                logger.info(f"No SJSU General applications parsed from {key}")
        except Exception as e:
            logger.error(f"Error processing {key}: {e}", exc_info=True)
            raise

    return {"statusCode": 200,
            "body": json.dumps({"message": f"Processed {records_written} SJSU General records"})}
